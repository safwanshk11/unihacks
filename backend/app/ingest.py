"""Read a raw catalogue file into RawProductIn rows.

Accepts CSV and XLSX, because the assessment hands over a dataset rather
than using the sample baked into this repo. Column matching is tolerant of
the header drift real spreadsheets carry (case, spaces vs underscores,
stray punctuation) and of the brief's warning that row 1 is not always the
header row.
"""

from __future__ import annotations

import csv
import io
import re

from app.models import RawProductIn

# Canonical field -> the header spellings seen in the wild for it.
COLUMN_ALIASES: dict[str, tuple[str, ...]] = {
    "mfg_part_num": ("mfgpartnum", "manufacturerpartnumber", "mfrpartnum", "partnumber", "mpn"),
    "part_desc": ("partdesc", "partdescription", "description", "itemdescription"),
    "e1_brand": ("e1brand",),
    "unilog_brand": ("unilogbrand",),
    "dib_brand": ("dibbrand",),
    "part_manuf": ("partmanuf", "partmanufacturer", "manufacturer", "manufacturername", "supplier"),
}

REQUIRED = ("mfg_part_num", "part_desc")


class IngestError(Exception):
    """Raised when a file can't be read as a product catalogue."""


def _norm(header: str) -> str:
    return re.sub(r"[^a-z0-9]", "", str(header or "").lower())


def _map_headers(headers: list[str]) -> dict[str, int]:
    """Canonical field name -> column index."""
    normalised = [_norm(h) for h in headers]
    mapping: dict[str, int] = {}
    for field, aliases in COLUMN_ALIASES.items():
        for candidate in (_norm(field), *aliases):
            if candidate in normalised:
                mapping[field] = normalised.index(candidate)
                break
    return mapping


def _find_header_row(rows: list[list], scan: int = 10) -> tuple[int, dict[str, int]]:
    """Spreadsheets from this pack carry title rows and merged cells above
    the real header, so locate the first row that actually maps."""
    best_idx, best_map = -1, {}
    for i, row in enumerate(rows[:scan]):
        mapping = _map_headers([str(c) if c is not None else "" for c in row])
        if len(mapping) > len(best_map):
            best_idx, best_map = i, mapping
        if all(f in mapping for f in REQUIRED) and len(mapping) >= 4:
            return i, mapping
    if best_idx >= 0 and all(f in best_map for f in REQUIRED):
        return best_idx, best_map
    raise IngestError(
        "Could not find a header row with at least 'Mfg_Part_Num' and 'Part_Desc'. "
        f"Looked at the first {scan} rows."
    )


def _rows_to_products(rows: list[list], header_idx: int, mapping: dict[str, int]) -> list[RawProductIn]:
    def cell(row: list, field: str) -> str:
        idx = mapping.get(field)
        if idx is None or idx >= len(row):
            return ""
        value = row[idx]
        return "" if value is None else str(value).strip()

    products: list[RawProductIn] = []
    for row in rows[header_idx + 1 :]:
        if not any(str(c).strip() for c in row if c is not None):
            continue
        mfg = cell(row, "mfg_part_num")
        desc = cell(row, "part_desc")
        if not mfg and not desc:
            continue
        products.append(
            RawProductIn(
                mfg_part_num=mfg or desc[:40],
                part_desc=desc or mfg,
                e1_brand=cell(row, "e1_brand") or None,
                unilog_brand=cell(row, "unilog_brand") or None,
                dib_brand=cell(row, "dib_brand") or None,
                part_manuf=cell(row, "part_manuf") or "",
            )
        )
    if not products:
        raise IngestError("No data rows found beneath the header row.")
    return products


def parse_catalog(filename: str, content: bytes) -> list[RawProductIn]:
    name = (filename or "").lower()

    if name.endswith((".xlsx", ".xlsm")):
        try:
            from openpyxl import load_workbook
        except ImportError as e:  # pragma: no cover
            raise IngestError("XLSX support requires openpyxl.") from e
        try:
            wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        except Exception as e:
            raise IngestError(f"Could not open the workbook: {e}") from e

        # The pack's own files put input and delivery data on separate
        # sheets, so pick the first sheet that actually looks like input.
        last_error: Exception | None = None
        for sheet in wb.worksheets:
            rows = [list(r) for r in sheet.iter_rows(values_only=True)]
            if not rows:
                continue
            try:
                header_idx, mapping = _find_header_row(rows)
                return _rows_to_products(rows, header_idx, mapping)
            except IngestError as e:
                last_error = e
        raise IngestError(f"No sheet looked like a product catalogue. {last_error or ''}".strip())

    if name.endswith(".csv") or not name:
        try:
            text = content.decode("utf-8-sig")
        except UnicodeDecodeError:
            text = content.decode("latin-1")
        rows = [list(r) for r in csv.reader(io.StringIO(text))]
        if not rows:
            raise IngestError("The file is empty.")
        header_idx, mapping = _find_header_row(rows)
        return _rows_to_products(rows, header_idx, mapping)

    raise IngestError(f"Unsupported file type '{filename}'. Upload a .csv or .xlsx file.")
